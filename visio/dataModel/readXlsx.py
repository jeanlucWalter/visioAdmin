from django import db
from openpyxl import load_workbook
from pathlib import Path
from unidecode import unidecode

class ReadXlsxRef:
  path = "visio/dataValues/xlsx/"
  __sheetName = "ReferentielVisio"
  __keyField = "Code SAP initial"
  __dicoFields = {
    "Code SAP Final": "PDV code",
    "Nom SAP": "PDV",
    "Libellé DRV SAP": "Drv",
    "Libellé Agent SAP": "Agent",
    "Dpt SAP" : "Département",
    "Libellé Bassin":"Bassin",
    "Ville SAP": "Ville",
    "Latitude": "Latitude",
    "Longitude": "Longitude",
    "Segmentation Portefeuille": "Segment Commercial",
    "Libellé Typologie SAP": "Segment Marketing",
    "Nom niveau 1": "Enseigne",
    "Nom niveau 2": "Ensemble",
    "Nom niveau 3": "Sous_Ensemble",
    "Nom niveau 4": "Site",
    "Point Feu": "Point Feu",
    # "Nb visites F"
    }

  def __init__(self, fileName, tablePdv):
    self.__tableIndex = tablePdv.json["tableIndex"]
    self.__dbValues = tablePdv.json['values']
    self.__titles = ["status"] + tablePdv.json['titles']
    self.__columnKey = None
    self.__columnField = {}
    self.__errors = []
    self.__dictValues = {}
    self.__isModified = False
    pathFile = Path.cwd() / f"{self.path}{fileName}.xlsx"
    if pathFile.exists():
        xlsxworkbook = load_workbook(pathFile, data_only=True)
        try:
          self.__sheet = xlsxworkbook[self.__sheetName]
        except:
          self.__errors.append(f"L'onglet {self.__sheetName} n'existe pas.")
        else:
          self.__readFields()
          self.__writeValues()
    else:
      self.__errors.append(f"Le fichier {fileName}.xlsx n'existe pas")

  @property
  def errors(self):
    return self.__errors

  @property
  def listValues(self) -> list:
    dictXlsx = dict(self.__dictValues)
    lastIndex = len(self.__dbValues[0]) - 1
    existingPdV = [self.__buildLine(dbaseLine[0], dictXlsx, dbaseLine, lastIndex) for dbaseLine in self.__dbValues]
    newPdv = [self.__lineBold(line) for line in dictXlsx.values()]
    return existingPdV + newPdv

  @property
  def json(self)->dict:
    return {'titles':self.__titles, 'values': self.listValues, 'tableIndex':self.__tableIndex}

  def __lineBold(self, line):
    return ['Nouveau'] + [f'<b>{value}</b>' for value in line] + ['<b>Nouvelle valeur</b>']

  def __buildLine(self, code:str, dictXlsx:dict, dbaseLine:list, lastIndex:int):
    if dbaseLine[lastIndex]:
      return ['Réouverture'] + self.__createNewValues(dictXlsx.pop(code), dbaseLine) + [dbaseLine[lastIndex]]
    if code in dictXlsx:
      newValues = self.__createNewValues(dictXlsx.pop(code), dbaseLine)
      status = 'Modifié' if self.__isModified else 'Normal'
      return [status] + newValues + ['']
    else:
      lastValue = dbaseLine[lastIndex] if dbaseLine[lastIndex] else 'Pdv fermé'
      return ['Fermeture'] + [f'<i>{val}</i>' for index, val in enumerate(dbaseLine) if index != lastIndex] + [f'<i>{lastValue}</i>']


  def __readFields(self):
    column, dictField = 1, {}
    value = self.__sheet.cell(row=1, column=column).value
    while value:
      if value == self.__keyField:
        self.__columnKey = column
      elif value in self.__dicoFields:
        dictField[value] = column
      column += 1
      value = self.__sheet.cell(row=1, column=column).value
    self.__columnField = [dictField[field] for field in self.__dicoFields.keys()]

  def __writeValues (self):
    row = 2
    while True:
      key = self.__sheet.cell(row=row, column=self.__columnKey).value
      if key:
        line = [self.__sheet.cell(row=row, column=column).value for column in self.__columnField]
        self.__dictValues[str(key)] = [value if value else "" for value in line]
        row += 1
      else:
        break

  def __createNewValues (self, codeLine, dbaseLine) -> list:
    self.__isModified = False
    return [self.__computeNewValue(codeItem, dbaseLine[index], index) for index, codeItem in enumerate(codeLine)]

  def __computeNewValue(self, codeItem, dbaseItem, index):
    if self.__computeNewValueTest(codeItem, dbaseItem, index):
      return dbaseItem
    self.__isModified = True
    return f"<i>{dbaseItem}</i> <b>{codeItem}</b>"

  def __computeNewValueTest(self, codeItem, dbaseItem, index):
    if codeItem == dbaseItem if type(codeItem) == type(dbaseItem) else str(codeItem) == str(dbaseItem):
      return True
    field = self.__titles[index + 1].strip()
    if field in ["Latitude", "Longitude"]:
      if type(dbaseItem) != float or type(codeItem) != float:
        return False
      return abs(float(codeItem) - float(dbaseItem)) < 0.0002
    elif field == "Enseigne": return self.__computEnseigneTest(codeItem, dbaseItem)
    elif field in ["Enseigne", "Ville", "Site", "Agent", "Bassin"]:
      return unidecode(codeItem.lower()).replace("-", " ") == unidecode(dbaseItem.lower()).replace("-", " ")
    elif field == "Point Feu":
      if codeItem == "N":
        return dbaseItem == ""
      if codeItem == "O":
        return dbaseItem == "Point Feu"
    elif field == "Segment Commercial":
      if dbaseItem == "non segmenté" and not codeItem:
        return True
      return dbaseItem != codeItem
    elif field == "Segment Marketing":
      if dbaseItem == "Autres métiers": 
        return codeItem in ["Autres métiers", "Plateforme de redist"]
    elif field in ["Ensemble", "Sous-Ensemble", "PDV"]:
      return codeItem.strip().replace("  ", " ") == dbaseItem.strip().replace("  ", " ")
    raise(TypeError)


  def __computEnseigneTest(self, codeItem, dbaseItem):
    if codeItem == "BIGMAT FRANCE":
      return dbaseItem == "CMEM"
    elif codeItem == "NEGOCES MATX DE CONSTRUCTION":
      return dbaseItem == "Nég Mtx Cons"
    elif codeItem == "POINT P - MATERIAUX DE CONSTRUCTION":
      return dbaseItem == "SGDB France"
    elif codeItem == "GROUPE(MENT)S REGIONAUX":
      return dbaseItem == "Group. Rég."
    elif codeItem == "BOIS & MATERIAUX":
      return dbaseItem == "Bois & Mat."
    elif codeItem == "NEGOCES AUTRES":
      return dbaseItem == "Autres"
    return unidecode(codeItem.lower()).replace("-", " ") == unidecode(dbaseItem.lower()).replace("-", " ")

class ReadXlsxVentes:
  path = "visio/dataValues/xlsx/"
  __sheetName = {"siniat":"Volumes Siniat", "salsi":"Volumes Salsi"}
  __keyField = {"siniat":["code pdv", 2], "salsi":["code pdv", 1]}
  __dicoFields = {
    "siniat":{"Plaques":"Siniat plaque",	"Cloisons":"Siniat cloison", "Doublages":"Siniat doublage",	"Enduit Prégy":"Pregy enduit",	"Mortier Prégy":"Pregy mortier"},
    "salsi": {"Enduit Joint Salsi":"Salsi enduit",	"Mortier Salsi":"Salsi mortier"}
  }

  def __init__(self, fileName, tableVentes):
    self.__tableIndex = tableVentes.json["tableIndex"]
    self.__dbValues = tableVentes.json['values']
    self.__titles = ["Status"] + tableVentes.json['titles']
    self.__errors = []
    self.__dictValues = {}
    pathFile = Path.cwd() / f"{self.path}{fileName}.xlsx"
    if pathFile.exists():
        xlsxworkbook = load_workbook(pathFile, data_only=True)
        try:
          self.__sheets = {key:xlsxworkbook[sheetName] for key, sheetName in self.__sheetName.items()}
        except:
          self.__errors.append(f"L'onglet {self.__sheetName} n'existe pas.")
        else:
          self.__dictField = self.__readFields()
          self.__dictValues = self.__readValues()
    else:
      self.__errors.append(f"Le fichier {fileName}.xlsx n'existe pas")

  @property
  def errors(self):
    return self.__errors

  @property
  def listValues(self) -> list:
    positionField = [key for key in self.__dictField["siniat"]] + [key for key in self.__dictField["salsi"]]
    return [self.__computeListValues(list, positionField) for list in self.__dbValues] + self.__missingInRef()

  @property
  def json(self)->dict:
    return {'titles':self.__titles, 'values': self.listValues, 'tableIndex':self.__tableIndex}

  def __computeListValues(self, list, positionField):
    pdvCode = list[0]
    if pdvCode in self.__dictValues:
      status = "Normal"
      for indexXlsx in range(len(positionField)):
        field = positionField[indexXlsx]
        indexBase = self.__titles.index(field) - 1
        if not list[indexBase]: list[indexBase] = "0.00"
        if not self.__dictValues[pdvCode][indexXlsx]: self.__dictValues[pdvCode][indexXlsx] = 0.0
        print("compute", list[indexBase], self.__dictValues[pdvCode][indexXlsx])
        if abs(float(list[indexBase].replace(" ", "").replace(",", ".")) - self.__dictValues[pdvCode][indexXlsx]) > 0.01:
          list[indexBase] = f'<i>{self.__formatNumber(list[indexBase])}</i> <b>{self.__formatNumber(self.__dictValues[pdvCode][indexXlsx])}</b>'
          status = "Modifié"
      return [status] + list
    return ["Absent du fichier Xls"] + [self.__printInRed(list[index]) if index > 4 else list[index] for index in range(len(list))]

  def __missingInRef(self):
    missing = []
    existingPdv = [listValues[0] for listValues in self.__dbValues]
    for pdvCode, listValues in self.__dictValues.items():
      # print(listValues)
      newList = list(listValues)
      if not pdvCode in existingPdv:
        newList.insert(3, "")
        # if all([True if type(element) == float else None for element in list]):
        #   newList[3] = 


        missing.append(["Absent du Référentiel", pdvCode, "", "", "", ""] + [self.__printInRed(self.__formatNumber(element)) for element in newList])
    return missing

    return missing

  def __printInRed(self, string):
    return '<span style="color:red">' + string + '</span>'

  def __formatNumber(self, value):
    if not value:
      value = "0.0"
    if type(value) == str:
      value = float(value.replace(" ", "").replace(",", "."))
    return f"{value:11,.2f}".replace(",", " ").replace(".", ",") if value != 0.0 else ""

  def __indexFields(self):
    indexField = {}
    positionField = [key for key in self.__dictField["siniat"]] + [key for key in self.__dictField["salsi"]]
    for field in self.__titles:
      if field in positionField:
        indexField[field] = positionField.index(field)
    return indexField

  def __readFields(self):
    column, dictField = 1, {}
    check = [sheet.cell(row=self.__keyField[key][1], column=column).value == self.__keyField[key][0] for key, sheet in self.__sheets.items()]
    if all(check):
      for key, sheet in self.__sheets.items():
        row, column = self.__keyField[key][1], 2
        dictField[key] = {}
        while True:
          fieldName = sheet.cell(row=row, column=column).value
          if fieldName in self.__dicoFields[key]:
            fieldForPrint = self.__dicoFields[key][fieldName]
            dictField[key][fieldForPrint] = column
          elif not fieldName:
            break
          column += 1
    return dictField

  def __readValues (self):
    dictValue = {}
    listValue = [None] * (len(self.__dictField["siniat"]) + len(self.__dictField["salsi"]))
    for key, sheet in self.__sheets.items():
      row = self.__keyField[key][1] + 1
      while True:
        cellValue = sheet.cell(row=row, column=1).value
        pdvCode = str(cellValue) if cellValue else False
        if pdvCode:
          if not pdvCode in dictValue:
            dictValue[pdvCode] = list(listValue)
          startIndex = 0 if key == "siniat" else len(self.__dictField["siniat"])
          index = 0
          for _, column in self.__dictField[key].items():
            dictValue[pdvCode][startIndex + index] = sheet.cell(row=row, column=column).value
            index += 1
        else:
          break
        row += 1
    return dictValue




